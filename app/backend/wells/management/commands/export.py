"""
    Licensed under the Apache License, Version 2.0 (the "License");
    you may not use this file except in compliance with the License.
    You may obtain a copy of the License at

        http://www.apache.org/licenses/LICENSE-2.0

    Unless required by applicable law or agreed to in writing, software
    distributed under the License is distributed on an "AS IS" BASIS,
    WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
    See the License for the specific language governing permissions and
    limitations under the License.
"""
import csv
import zipfile
import os
import logging
import string

from django.core.management.base import BaseCommand
from django.db import models
from django.db import connection

from minio import Minio
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.worksheet.write_only import WriteOnlyCell

from gwells.settings.base import get_env_variable
from gwells.management.commands import ResultIter

# Run from command line :
# python manage.py export
#
# For development/debugging, it's useful to skip upload and cleanup
# python manage.py export --cleanup=0 --upload=0

logger = logging.getLogger(__name__)


class Command(BaseCommand):

    def add_arguments(self, parser):
        # Arguments added for debugging purposes.
        # e.g. don't cleanup, don't upload: python manage.py export_databc --cleanup=0 --upload=0
        parser.add_argument('--cleanup', type=int, nargs='?', help='If 1, remove file when done', default=1)
        parser.add_argument('--upload', type=int, nargs='?', help='If 1, upload the file', default=1)

    def handle(self, *args, **options):
        logger.info('starting export')
        zip_filename = 'gwells.zip'
        spreadsheet_filename = 'gwells.xlsx'
        self.generate_files(zip_filename, spreadsheet_filename)
        if options['upload'] == 1:
            self.upload_files(zip_filename, spreadsheet_filename)
        if options['cleanup'] == 1:
            logger.info('cleaning up')
            for filename in (zip_filename, spreadsheet_filename):
                if os.path.exists(filename):
                    os.remove(filename)
        logger.info('export complete')
        self.stdout.write(self.style.SUCCESS('export complete'))

    def upload_files(self, zip_filename, spreadsheet_filename):
        is_secure = get_env_variable('S3_USE_SECURE', '1', warn=False) is '1'
        minioClient = Minio(get_env_variable('S3_HOST'),
                            access_key=get_env_variable('S3_PUBLIC_ACCESS_KEY'),
                            secret_key=get_env_variable('S3_PUBLIC_SECRET_KEY'),
                            secure=is_secure)
        for filename in (zip_filename, spreadsheet_filename):
            logger.info('uploading {}'.format(filename))
            with open(filename, 'rb') as file_data:
                file_stat = os.stat(filename)
                # Do we need to remove the existing files 1st?
                target = 'export/{}'.format(filename)
                minioClient.put_object(get_env_variable('S3_WELL_EXPORT_BUCKET'),
                                       target,
                                       file_data,
                                       file_stat.st_size)

    def export(self, workbook, gwells_zip, worksheet_name, cursor):
        logger.info('exporting {}'.format(worksheet_name))
        worksheet = workbook.create_sheet(worksheet_name)
        csv_file = '{}.csv'.format(worksheet_name)
        # If any of the export files already exist, delete them
        if os.path.exists(csv_file):
            os.remove(csv_file)
        with open(csv_file, 'w') as csvfile:
            csvwriter = csv.writer(csvfile, dialect='excel')

            values = []
            cells = []
            # Write the headings
            for index, field in enumerate(cursor.description):
                fieldName = field[0]
                values.append(fieldName)
                cell = WriteOnlyCell(worksheet, value=fieldName)
                cell.font = Font(bold=True)
                cells.append(cell)
            columns = len(values)

            for index, value in enumerate(values):
                worksheet.column_dimensions[get_column_letter(index+1)].width = len(value) + 2

            worksheet.append(cells)
            csvwriter.writerow(values)

            # Write the values
            row_index = 0
            for row, record in enumerate(ResultIter(cursor)):
                values = []
                num_values = 0
                for col, value in enumerate(record):
                    if not (value == "" or value is None):
                        num_values += 1
                    if type(value) is str:
                        # There are lots of non-printable characters in the source data that can cause
                        # issues in the export, so we have to clear them out.
                        v = ''.join([s for s in value if s in string.printable])
                        # We can't have something starting with an = sign,
                        # it would be interpreted as a formula in excel.
                        if v.startswith('='):
                            v = '\'{}'.format(v)
                        values.append(v)
                    else:
                        values.append(value)
                if num_values > 1:
                    # We always have a well_tag_number, but if that's all we have, then just skip this record
                    row_index += 1
                    csvwriter.writerow(values)
                    worksheet.append(values)

            filter_reference = 'A1:{}{}'.format(get_column_letter(columns), row_index+1)
            worksheet.auto_filter.ref = filter_reference

        gwells_zip.write(csv_file)
        if os.path.exists(csv_file):
            # After adding the csv file to the zip, delete it.
            os.remove(csv_file)

    def generate_files(self, zip_filename, spreadsheet_filename):
        #######
        # WELL
        #######
        well_sql = ("""select well_tag_number, identification_plate_number,
 well_identification_plate_attached,
 well_status_code, well.well_class_code,
 wsc.well_subclass_code as well_subclass,
 intended_water_use_code, licenced_status_code,
 observation_well_number, obs_well_status_code, water_supply_system_name,
 water_supply_system_well_name,
 well.street_address, well.city, legal_lot, legal_plan, legal_district_lot, legal_block,
 legal_section, legal_township, legal_range,
 land_district_code,
 legal_pid,
 well_location_description,
 st_y(geom) as latitude, st_x(geom) as longitude, utm_zone_code, utm_northing, utm_easting,
 coordinate_acquisition_code, bcgs_id,
 construction_start_date, construction_end_date, alteration_start_date,
 alteration_end_date, decommission_start_date, decommission_end_date,
 driller_name, consultant_name, consultant_company,
 diameter, total_depth_drilled, finished_well_depth, final_casing_stick_up,
 bedrock_depth, ground_elevation, ground_elevation_method_code, static_water_level,
 well_yield,
 well_yield_unit_code,
 artesian_flow, artesian_pressure, well_cap_type, well_disinfected_code,
 well_orientation_code,
 alternative_specs_submitted,
 surface_seal_material_code, surface_seal_method_code, surface_seal_length, surface_seal_depth,
 backfill_type,
 backfill_depth,
 liner_material_code, liner_diameter, liner_thickness, surface_seal_thickness,
 liner_from, liner_to,
 screen_intake_method_code, screen_type_code, screen_material_code,
 other_screen_material, screen_information,
 screen_opening_code, screen_bottom_code, other_screen_bottom,
 filter_pack_from,
 filter_pack_to, filter_pack_material_code,
 filter_pack_thickness,
 filter_pack_material_size_code,
 development_hours, development_notes,
 water_quality_colour, water_quality_odour,
 yield_estimation_method_code,
 yield_estimation_rate,
 yield_estimation_duration, static_level_before_test, drawdown,
 hydro_fracturing_performed, hydro_fracturing_yield_increase,
 decommission_reason, decommission_method_code, decommission_details, decommission_sealant_material,
 decommission_backfill_material,
 comments,
 ems,
 registries_person.surname as person_responsible,
 registries_organization.name as company_of_person_responsible,
 aquifer_id,
 aquifer_vulnerability_index as avi,
 storativity,
 transmissivity,
 hydraulic_conductivity,
 specific_storage,
 specific_yield,
 testing_method,
 testing_duration,
 analytic_solution_type,
 boundary_effect_code,
 aquifer_lithology_code
 from well
 left join well_subclass_code as wsc on wsc.well_subclass_guid = well.well_subclass_guid
 left join registries_person on
 registries_person.person_guid = well.person_responsible_guid
 left join registries_organization on
 registries_organization.org_guid = well.org_of_person_responsible_guid
 where well.well_publication_status_code = 'Published' or well.well_publication_status_code = null
 order by well_tag_number""")
        ###########
        # LITHOLOGY
        ###########
        lithology_sql = ("""select lithology_description.well_tag_number, lithology_from, lithology_to,
 lithology_raw_data,
 ldc.description as lithology_description_code,
 lmc.description as lithology_material_code,
 lhc.description as lithology_hardness_code,
 lcc.description as lithology_colour_code,
 water_bearing_estimated_flow,
 lithology_description.well_yield_unit_code, lithology_observation
 from lithology_description
 left join lithology_description_code as ldc on
 ldc.lithology_description_code = lithology_description.lithology_description_code
 left join lithology_material_code as lmc on
 lmc.lithology_material_code = lithology_description.lithology_material_code
 left join lithology_hardness_code as lhc on
 lhc.lithology_hardness_code = lithology_description.lithology_hardness_code
 left join lithology_colour_code as lcc on
 lcc.lithology_colour_code = lithology_description.lithology_colour_code
 inner join well on well.well_tag_number = lithology_description.well_tag_number
 where well.well_publication_status_code = 'Published' or well.well_publication_status_code = null
 order by lithology_description.well_tag_number""")
        ########
        # CASING
        ########
        casing_sql = ("""select casing.well_tag_number, casing_from, casing_to, casing.diameter, casing_code,
 casing_material_code, wall_thickness, drive_shoe_code
 from casing
 inner join well on well.well_tag_number = casing.well_tag_number
 where well.well_publication_status_code = 'Published' or well.well_publication_status_code = null
 order by casing.well_tag_number""")
        ########
        # SCREEN
        ########
        screen_sql = ("""select screen.well_tag_number, screen_from, screen_to, screen_diameter,
 screen_assembly_type_code, slot_size
 from screen
 inner join well on well.well_tag_number = screen.well_tag_number
 where well.well_publication_status_code = 'Published' or well.well_publication_status_code = null
 order by screen.well_tag_number""")
        ##############
        # PERFORATIONS
        ##############
        perforation_sql = ("""select p.well_tag_number, p.liner_perforation_from, p.liner_perforation_to
 from liner_perforation as p
 inner join well on well.well_tag_number = p.well_tag_number
 where well.well_publication_status_code = 'Published' or well.well_publication_status_code = null
 order by p.well_tag_number""")
        #################
        # DRILLING METHOD
        #################
        drilling_method_sql = ("""select well_id as well_tag_number,
drillingmethodcode_id as drilling_method_code
from well_drilling_methods
inner join well on well.well_tag_number = well_drilling_methods.well_id
where well.well_publication_status_code = 'Published' or well.well_publication_status_code = null
order by well_tag_number""")
        ####################
        # DEVELOPMENT METHOD
        ####################
        development_method_sql = ("""select well_id as well_tag_number,
developmentmethodcode_id as development_method_code
from well_development_methods
inner join well on well.well_tag_number = well_development_methods.well_id
where well.well_publication_status_code = 'Published' or well.well_publication_status_code = null
order by well_tag_number""")

        # Create a dictionary to iterate through when creating spreadsheets.
        sheets = {
            'well': well_sql,
            'lithology': lithology_sql,
            'casing': casing_sql,
            'screen': screen_sql,
            'perforation': perforation_sql,
            'drilling_method': drilling_method_sql,
            'development_method': development_method_sql
        }

        # If there is an existing zip file, remove it.
        if os.path.exists(zip_filename):
            os.remove(zip_filename)
        with zipfile.ZipFile(zip_filename, 'w', compression=zipfile.ZIP_DEFLATED) as gwells_zip:
            if os.path.exists(spreadsheet_filename):
                os.remove(spreadsheet_filename)
            workbook = Workbook(write_only=True)

            for sheet, sql in sheets.items():
                logger.info('creating {} cursor'.format(sheet))
                with connection.cursor() as cursor:
                    cursor.execute(sql)
                    self.export(workbook, gwells_zip, sheet, cursor)
            workbook.save(filename=spreadsheet_filename)
            # Add a readme to the zipfile.
            arcname = 'README.md'
            readme = os.path.join(os.path.dirname(__file__), arcname)
            gwells_zip.write(readme, arcname)
